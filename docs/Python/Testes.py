
import time
from selenium import webdriver
import pandas as pd
import openpyxl
import xlsxwriter
import datetime


diretorio="C:\\Users\\Padulla\\Documents\\GitHub\\corona_dash\\docs\\Python\\"


data=datetime.date.today().strftime('%d/%m/%Y')



driver = webdriver.Chrome('F:/chromedriver')  # Optional argument, if not specified will search path.

driver.get('https://www.covid-19.pa.gov.br/#/');
time.sleep(3) # Let the user actually see something! 
driver.execute_script("window.scrollTo(0, 4200)") 
time.sleep(10) # Let the user actually see something!
leito_clinico_total = driver.find_elements_by_xpath('//*[@id="q-app"]/div/div/div[8]/div[2]/div/div/div/div/div[2]/div[2]/div/div/table/tbody/tr[1]/td[2]/span')[0]
leito_clinico_disp = driver.find_elements_by_xpath('//*[@id="q-app"]/div/div/div[8]/div[2]/div/div/div/div/div[2]/div[2]/div/div/table/tbody/tr[1]/td[3]/span')[0]
leito_clinico_ocup  = driver.find_elements_by_xpath('//*[@id="q-app"]/div/div/div[8]/div[2]/div/div/div/div/div[2]/div[2]/div/div/table/tbody/tr[1]/td[4]/span')[0]
leito_uti_total = driver.find_elements_by_xpath('//*[@id="q-app"]/div/div/div[8]/div[2]/div/div/div/div/div[2]/div[2]/div/div/table/tbody/tr[3]/td[2]/span')[0]
leito_uti_disp = driver.find_elements_by_xpath('//*[@id="q-app"]/div/div/div[8]/div[2]/div/div/div/div/div[2]/div[2]/div/div/table/tbody/tr[3]/td[3]/span')[0]
leito_uti_ocup = driver.find_elements_by_xpath('//*[@id="q-app"]/div/div/div[8]/div[2]/div/div/div/div/div[2]/div[2]/div/div/table/tbody/tr[3]/td[4]/span')[0]


leito_clinico_total = int(leito_clinico_total.text)
leito_clinico_disp = int(leito_clinico_disp.text)
leito_clinico_ocup = leito_clinico_ocup.text
leito_uti_total = int(leito_uti_total.text)
leito_uti_disp = int(leito_uti_disp.text)
leito_uti_ocup = leito_uti_ocup.text



driver.quit()
dados=[data,leito_clinico_total,leito_clinico_disp,leito_clinico_ocup,leito_uti_total,leito_uti_disp,leito_uti_ocup]


wb = openpyxl.load_workbook(filename=diretorio+'PA.xlsx')
sheet=wb.sheetnames
sheet_1=wb[wb.sheetnames[0]]
row=sheet_1.max_row+1


#Data
sheet_1 .cell(row = row, column = 1).value = dados[0]
sheet_1 .cell(row = row, column = 2).value = dados[1]
sheet_1 .cell(row = row, column = 3).value = dados[2]
sheet_1 .cell(row = row, column = 4).value = dados[3]
sheet_1 .cell(row = row, column = 5).value = dados[4]
sheet_1 .cell(row = row, column = 6).value = dados[5]
sheet_1 .cell(row = row, column = 7).value = dados[6]

wb.save(filename=diretorio+'PA.xlsx')





